"""
模块五：选品方向建议（基于历史销售和库存）

用法:
  python scripts/product_strategy.py
  python scripts/product_strategy.py --csv-dir ./data
  python scripts/product_strategy.py --api-mode

输入:
  - sales_90d.csv: 近90天销售明细
  - inventory.csv: 当前库存快照
输出:
  - product_strategy.xlsx (4个Sheet: 明星品/现金牛/问题品/淘汰品)
"""

import argparse
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

_project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_project_root))

from shared.config import load_config, PROJECT_ROOT
from shared.logger import setup_logger
from shared.data_source import create_data_source
from shared.models import to_usd

logger = setup_logger("product_strategy")


def load_sales_90d(csv_path: str) -> list[dict]:
    """加载90天销售数据"""
    p = Path(csv_path)
    if not p.exists():
        logger.error(f"销售数据文件不存在: {csv_path}")
        sys.exit(1)
    import csv
    with open(p, "r", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def load_inventory_csv(csv_path: str) -> list[dict]:
    """加载库存数据"""
    p = Path(csv_path)
    if not p.exists():
        logger.warning(f"库存文件不存在: {csv_path}")
        return []
    import csv
    with open(p, "r", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def classify_products(sales_rows: list[dict], inventory_rows: list[dict],
                      replenish_lead_time: int = 40) -> dict:
    """对SKU做四象限分类"""
    # 按SKU聚合销售数据
    sku_sales: dict[str, dict] = defaultdict(lambda: {
        "sku": "", "client_sku": "", "qty_90d": 0,
        "revenue_90d": 0.0, "cost_90d": 0.0,
        "last_sale_date": None, "qty_30d": 0,
    })

    now = datetime.now()
    cutoff_90d = now - timedelta(days=90)
    cutoff_30d = now - timedelta(days=30)

    for row in sales_rows:
        sku = row.get("SKU") or row.get("sku") or ""
        client_sku = row.get("ClientSKU") or row.get("client_sku") or ""
        if not sku:
            continue

        qty = int(row.get("数量") or row.get("qty") or row.get("product_num", 0))
        sales = float(row.get("销售额") or row.get("sales") or row.get("revenue", 0))
        cost = float(row.get("成本") or row.get("cost", 0))

        try:
            sale_date = datetime.strptime(
                row.get("日期") or row.get("date") or row.get("order_date", ""),
                "%Y-%m-%d"
            )
        except (ValueError, TypeError):
            sale_date = None

        d = sku_sales[sku]
        d["sku"] = sku
        d["client_sku"] = client_sku or d["client_sku"]
        d["qty_90d"] += qty
        d["revenue_90d"] += sales
        d["cost_90d"] += cost * qty

        if sale_date:
            if d["last_sale_date"] is None or sale_date > d["last_sale_date"]:
                d["last_sale_date"] = sale_date

        if sale_date and sale_date >= cutoff_30d:
            d["qty_30d"] += qty

    # 构建库存字典
    inv_map = {}
    for row in inventory_rows:
        sku = row.get("SKU") or row.get("sku") or ""
        try:
            inv_map[sku] = {
                "quantity": int(row.get("quantity") or row.get("库存", 0)),
                "safety_stock": int(row.get("safety_stock") or row.get("安全库存", 0)),
                "in_transit": int(row.get("in_transit") or row.get("在途库存", 0)),
            }
        except (ValueError, TypeError):
            pass

    # 分类
    star = []
    cash_cow = []
    question = []
    eliminate = []

    for sku, d in sku_sales.items():
        inv = inv_map.get(sku, {"quantity": 0, "safety_stock": 0, "in_transit": 0})
        stock_qty = inv["quantity"]
        safety = inv["safety_stock"] or 30  # 默认安全库存30天
        in_transit = inv.get("in_transit", 0)

        # 毛利率
        gross_margin = (d["revenue_90d"] - d["cost_90d"]) / d["revenue_90d"] * 100 if d["revenue_90d"] > 0 else 0.0

        # 日均销量
        avg_daily_sales = d["qty_90d"] / 90

        # 库存天数（日均销量为0时视为无限大）
        stock_days = stock_qty / avg_daily_sales if avg_daily_sales > 0 else 999

        # 近30天是否有销售
        sold_recently = d["qty_30d"] > 0

        item = {
            "SKU": sku,
            "ClientSKU": d["client_sku"],
            "90天销量": d["qty_90d"],
            "90天销售额": round(d["revenue_90d"], 2),
            "毛利率": round(gross_margin, 1),
            "当前库存": stock_qty,
            "在途库存": in_transit,
            "安全库存": safety,
            "日均销量(90天)": round(avg_daily_sales, 2),
            "库存天数": round(stock_days, 0),
            "最后销售日期": d["last_sale_date"].strftime("%Y-%m-%d") if d["last_sale_date"] else "无",
            "近30天销售": d["qty_30d"],
        }

        # 分类逻辑
        if sold_recently and d["qty_90d"] >= 10 and gross_margin > 30 and stock_days < replenish_lead_time:
            # 明星品：有销量、高毛利、库存快见底
            d["category"] = "star"
            item["建议补货量"] = max(0, int(avg_daily_sales * (replenish_lead_time + safety / 30 * replenish_lead_time) - stock_qty - in_transit))
            item["建议补货量2"] = max(0, int(avg_daily_sales * 30 * 1.5 - stock_qty))
            star.append(item)
        elif sold_recently and d["qty_90d"] >= 10 and 15 <= gross_margin <= 30 and 30 <= stock_days <= 60:
            # 现金牛：稳定销量、中等毛利、库存合理
            cash_cow.append(item)
        elif (d["qty_90d"] < 10 or gross_margin < 15 or stock_days > 90) and sold_recently:
            # 问题品：销量低、毛利低、或库存积压
            item["问题分析"] = ""
            if d["qty_90d"] < 10:
                item["问题分析"] += "销量过低; "
            if gross_margin < 15:
                item["问题分析"] += "毛利率偏低; "
            if stock_days > 90:
                item["问题分析"] += f"库存积压({stock_days:.0f}天); "
            # 行动建议
            actions = []
            if stock_days > 90:
                actions.append("降价清仓/捆绑销售")
            if gross_margin < 15:
                actions.append("优化采购成本/提价")
            if d["qty_90d"] < 10:
                actions.append("检查Listing质量/暂停广告")
            item["行动建议"] = " | ".join(actions) if actions else "重新评估产品定位"
            question.append(item)
        elif not sold_recently and gross_margin < 0:
            # 淘汰品：近期无销售且毛利率为负
            eliminate.append(item)
        elif not sold_recently:
            # 近30天无销售，归为问题品
            item["问题分析"] = "近30天无销售; "
            item["行动建议"] = "检查Listing/考虑下架或清仓"
            question.append(item)
        else:
            # 其他归为问题品
            item["问题分析"] = "不符合明星/现金牛标准"
            question.append(item)

    # 排序
    star.sort(key=lambda x: x["90天销售额"], reverse=True)
    cash_cow.sort(key=lambda x: x["90天销售额"], reverse=True)
    question.sort(key=lambda x: x["90天销量"], reverse=True)
    eliminate.sort(key=lambda x: x["90天销量"])

    logger.info(f"分类结果: 明星品{len(star)} | 现金牛{len(cash_cow)} | 问题品{len(question)} | 淘汰品{len(eliminate)}")
    return {"star": star, "cash_cow": cash_cow, "question": question, "eliminate": eliminate}


def export_to_excel(categories: dict, output_path: str):
    """输出四Sheet Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.error("需要 openpyxl: pip install openpyxl")
        csv_path = output_path.replace(".xlsx", ".csv")
        all_rows = []
        for cat_name, items in categories.items():
            for item in items:
                item["分类"] = cat_name
                all_rows.append(item)
        import csv
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            if all_rows:
                w = csv.DictWriter(f, fieldnames=list(all_rows[0].keys()))
                w.writeheader()
                w.writerows(all_rows)
        logger.info(f"已降级输出CSV: {csv_path}")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    category_names = {
        "star": ("明星品", "00B050"),
        "cash_cow": ("现金牛", "2F75B5"),
        "question": ("问题品", "FFC000"),
        "eliminate": ("淘汰品", "C00000"),
    }

    for cat_key, (cat_display, color_hex) in category_names.items():
        items = categories.get(cat_key, [])
        ws = wb.create_sheet(cat_display)

        if not items:
            ws.cell(row=1, column=1, value="无数据")
            continue

        headers = list(items[0].keys())
        header_fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        red_font = Font(color="FF0000")
        for row_idx, row_data in enumerate(items, 2):
            for col_idx, h in enumerate(headers, 1):
                val = row_data.get(h, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if isinstance(val, (int, float)) and val < 0:
                    cell.font = red_font

        ws.column_dimensions["A"].width = 16
        ws.freeze_panes = "A2"

    wb.save(output_path)
    logger.info(f"选品策略报告已保存: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="赛盒ERP选品策略分析")
    parser.add_argument("--csv-dir", type=str, help="CSV数据目录")
    parser.add_argument("--api-mode", action="store_true", help="使用API(实验性)")
    parser.add_argument("--config", type=str, help="配置文件路径")
    parser.add_argument("--output", type=str, default=None, help="输出文件路径")
    parser.add_argument("--lead-time", type=int, default=40,
                        help="补货提前期(天)，默认40天(海运)")
    args = parser.parse_args()

    config = load_config(args.config)
    if args.csv_dir:
        config.csv_dir = args.csv_dir
        config.resolve_csv_paths()
    config.resolve_csv_paths()

    logger.info(f"选品策略分析开始 (补货提前期={args.lead_time}天)")

    # 加载数据
    if args.api_mode:
        logger.info("API模式: 从订单数据提取销售信息")
        ds = create_data_source(config)
        orders = ds.fetch_orders(days=90)
        # 转换为销售行格式
        sales_rows = []
        for order in orders:
            for item in order.items:
                sales_rows.append({
                    "sku": item.sku,
                    "client_sku": item.client_sku,
                    "qty": item.qty,
                    "revenue": item.revenue_usd,
                    "cost": item.cost,
                })
        inventory_raw = ds.fetch_inventory()
        inventory_rows = [
            {"sku": i.sku, "quantity": i.quantity,
             "safety_stock": i.safety_stock, "in_transit": i.in_transit}
            for i in inventory_raw
        ]
        logger.info(f"从API提取: {len(sales_rows)} 行销售, {len(inventory_rows)} 库存")
    else:
        sales_rows = load_sales_90d(config.sales_90d_csv)
        inventory_rows = load_inventory_csv(config.inventory_csv)

    if not sales_rows:
        logger.error("销售数据为空")
        sys.exit(1)

    # 四象限分类
    categories = classify_products(sales_rows, inventory_rows, replenish_lead_time=args.lead_time)

    # 打印摘要
    print("\n=== 选品策略分类结果 ===")
    for cat_key, cat_display in [("star", "明星品"), ("cash_cow", "现金牛"),
                                  ("question", "问题品"), ("eliminate", "淘汰品")]:
        items = categories.get(cat_key, [])
        total_rev = sum(i.get("90天销售额", 0) for i in items)
        print(f"  {cat_display}: {len(items)} 个SKU, 90天总销售额 ${total_rev:.2f}")

    if categories["star"]:
        print("\n=== 明星品补货建议 ===")
        for item in categories["star"][:10]:
            print(f"  {item['ClientSKU'] or item['SKU']}: "
                  f"当前库存{item['当前库存']}, "
                  f"建议补货{item.get('建议补货量', 'N/A')}")

    # 输出Excel
    output_dir = PROJECT_ROOT / "reports"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = args.output or str(output_dir / f"product_strategy_{datetime.now().strftime('%Y-%m-%d')}.xlsx")
    export_to_excel(categories, output_path)

    logger.info("选品策略分析完成")


if __name__ == "__main__":
    main()
