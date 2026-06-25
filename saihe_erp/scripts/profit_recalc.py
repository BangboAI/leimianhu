"""
模块二：自定义利润重算（按运营人员分摊）

用法:
  python scripts/profit_recalc.py
  python scripts/profit_recalc.py --csv-dir ./data
  python scripts/profit_recalc.py --api-mode

输入:
  - profit_detail.csv: 利润明细（从赛盒导出）
  - operator_mapping.csv: SKU→运营人映射
输出:
  - profit_by_operator.xlsx (汇总+明细两个sheet)
"""

import argparse
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Optional

_project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_project_root))

from shared.config import load_config, PROJECT_ROOT
from shared.logger import setup_logger
from shared.data_source import CSVDataSource

logger = setup_logger("profit_recalc")


def load_profit_data(csv_path: str) -> list[dict]:
    """加载利润明细CSV"""
    p = Path(csv_path)
    if not p.exists():
        logger.error(f"利润明细文件不存在: {csv_path}")
        sys.exit(1)
    import csv
    with open(p, "r", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def load_operator_mapping(csv_path: str) -> dict[str, str]:
    """加载运营人映射，返回 {sku: operator}"""
    p = Path(csv_path)
    if not p.exists():
        logger.warning(f"运营人映射文件不存在: {csv_path}，将使用'未分配'")
        return {}
    import csv
    mapping = {}
    with open(p, "r", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            sku = (row.get("sku") or row.get("client_sku") or "").strip()
            operator = (row.get("operator") or "").strip()
            if sku and operator:
                mapping[sku] = operator
    logger.info(f"加载了 {len(mapping)} 条运营人映射")
    return mapping


def compute_by_operator(rows: list[dict], op_map: dict[str, str]) -> dict:
    """按运营人聚合利润"""
    op_data = defaultdict(lambda: {
        "operator": "",
        "total_sales": 0.0, "total_refund": 0.0,
        "total_cost": 0.0, "total_shipping": 0.0,
        "total_platform_fee": 0.0, "total_ad_spend": 0.0,
        "total_other": 0.0, "net_profit": 0.0,
        "sku_count": set(), "total_profit": 0.0,
    })
    sku_details = []

    for row in rows:
        sku = row.get("SKU") or row.get("sku") or ""
        client_sku = row.get("ClientSKU") or row.get("client_sku") or ""
        store = row.get("店铺") or row.get("store") or ""
        sales = float(row.get("销售额") or row.get("sales") or 0)
        refund = float(row.get("退款") or row.get("refund") or 0)
        cost = float(row.get("采购成本") or row.get("cost") or 0)
        shipping = float(row.get("头程运费") or row.get("shipping") or 0)
        plat_fee = float(row.get("平台费用") or row.get("platform_fee") or 0)
        ad = float(row.get("广告费") or row.get("ad_spend") or 0)
        other = float(row.get("其他费用") or row.get("other") or 0)

        identifier = client_sku or sku
        operator = op_map.get(sku) or op_map.get(client_sku) or "未分配"

        net = sales - refund - cost - shipping - plat_fee - ad - other
        margin_pct = (net / sales * 100) if sales > 0 else 0.0

        op = op_data[operator]
        op["operator"] = operator
        op["total_sales"] += sales
        op["total_refund"] += refund
        op["total_cost"] += cost
        op["total_shipping"] += shipping
        op["total_platform_fee"] += plat_fee
        op["total_ad_spend"] += ad
        op["total_other"] += other
        op["net_profit"] += net
        op["sku_count"].add(identifier)

        sku_details.append({
            "SKU": sku,
            "ClientSKU": client_sku,
            "店铺": store,
            "运营人": operator,
            "销售额": round(sales, 2),
            "退款": round(refund, 2),
            "采购成本": round(cost, 2),
            "头程运费": round(shipping, 2),
            "平台费用": round(plat_fee, 2),
            "广告费": round(ad, 2),
            "其他费用": round(other, 2),
            "净利润": round(net, 2),
            "利润率": round(margin_pct, 1),
        })

    # 计算每人衍生指标
    op_summary = []
    for op_name, d in sorted(op_data.items(), key=lambda x: x[1]["net_profit"], reverse=True):
        sku_count = len(d["sku_count"])
        avg_profit_per_sku = d["net_profit"] / sku_count if sku_count > 0 else 0.0
        margin = d["net_profit"] / d["total_sales"] * 100 if d["total_sales"] > 0 else 0.0
        op_summary.append({
            "运营人": op_name,
            "总销售额": round(d["total_sales"], 2),
            "总退款": round(d["total_refund"], 2),
            "采购成本": round(d["total_cost"], 2),
            "头程运费": round(d["total_shipping"], 2),
            "平台费用": round(d["total_platform_fee"], 2),
            "广告费": round(d["total_ad_spend"], 2),
            "其他费用": round(d["total_other"], 2),
            "净利润": round(d["net_profit"], 2),
            "利润率": f"{margin:.1f}%",
            "管理SKU数": sku_count,
            "平均单品利润": round(avg_profit_per_sku, 2),
        })

    return op_summary, sku_details


def export_to_excel(op_summary: list[dict], sku_details: list[dict], output_path: str):
    """输出双Sheet Excel文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
    except ImportError:
        logger.error("需要 openpyxl: pip install openpyxl")
        # 降级到CSV
        csv_path = output_path.replace(".xlsx", "_汇总.csv")
        import csv
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            if op_summary:
                w = csv.DictWriter(f, fieldnames=op_summary[0].keys())
                w.writeheader()
                w.writerows(op_summary)
        csv_path2 = output_path.replace(".xlsx", "_明细.csv")
        with open(csv_path2, "w", newline="", encoding="utf-8-sig") as f:
            if sku_details:
                w = csv.DictWriter(f, fieldnames=sku_details[0].keys())
                w.writeheader()
                w.writerows(sku_details)
        logger.info(f"已降级输出CSV: {csv_path}, {csv_path2}")
        return

    wb = openpyxl.Workbook()
    red_font = Font(color="FF0000")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Sheet 1: 运营人汇总
    ws1 = wb.active
    ws1.title = "运营人汇总"
    if op_summary:
        headers = list(op_summary[0].keys())
        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row_data in enumerate(op_summary, 2):
            for col_idx, h in enumerate(headers, 1):
                val = row_data[h]
                cell = ws1.cell(row=row_idx, column=col_idx, value=val)
                if isinstance(val, (int, float)) and val < 0:
                    cell.font = red_font

        ws1.column_dimensions["A"].width = 16
        for col in "BCDEFGHIJ":
            ws1.column_dimensions[col].width = 14
        ws1.freeze_panes = "A2"

    # Sheet 2: SKU明细
    ws2 = wb.create_sheet("SKU明细")
    if sku_details:
        headers = list(sku_details[0].keys())
        for col, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row_data in enumerate(sku_details, 2):
            for col_idx, h in enumerate(headers, 1):
                val = row_data[h]
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                if isinstance(val, (int, float)) and val < 0:
                    cell.font = red_font

        for col in "ABCDEFGHIJ":
            ws2.column_dimensions[col].width = 14
        ws2.freeze_panes = "A2"

    wb.save(output_path)
    logger.info(f"Excel已保存: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="赛盒ERP利润重算")
    parser.add_argument("--csv-dir", type=str, help="CSV数据目录")
    parser.add_argument("--api-mode", action="store_true", help="从API拉取数据(实验性)")
    parser.add_argument("--config", type=str, help="配置文件路径")
    parser.add_argument("--output", type=str, default=None, help="输出文件路径")
    args = parser.parse_args()

    config = load_config(args.config)
    if args.csv_dir:
        config.csv_dir = args.csv_dir
        config.resolve_csv_paths()
    if args.api_mode:
        config.data_mode = "api"

    config.resolve_csv_paths()
    logger.info(f"利润重算 - CSV目录: {config.csv_dir}")

    # 加载数据
    rows = load_profit_data(config.profit_detail_csv)
    op_map = load_operator_mapping(config.operator_mapping_csv)

    if not rows:
        logger.error("利润明细数据为空")
        sys.exit(1)

    # 计算
    op_summary, sku_details = compute_by_operator(rows, op_map)

    # 打印摘要
    logger.info(f"共 {len(rows)} 条记录, {len(op_summary)} 位运营人")
    print("\n=== 运营人利润汇总 ===")
    print(f"{'运营人':<12} {'销售额':>12} {'净利润':>12} {'利润率':>10} {'SKU数':>8} {'单品利润':>10}")
    print("-" * 64)
    for op in op_summary:
        print(f"{op['运营人']:<12} ${op['总销售额']:>9.2f} ${op['净利润']:>9.2f} "
              f"{op['利润率']:>10} {op['管理SKU数']:>8} ${op['平均单品利润']:>8.2f}")

    # 输出
    output_dir = PROJECT_ROOT / "reports"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = args.output or str(output_dir / f"profit_by_operator_{datetime.now().strftime('%Y-%m-%d')}.xlsx")
    export_to_excel(op_summary, sku_details, output_path)

    logger.info("利润重算完成")


if __name__ == "__main__":
    main()
